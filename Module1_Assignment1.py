{
 "cells": [
  {
   "cell_type": "code",
   "execution_count": 1,
   "id": "15a234af-a437-4f7e-81f7-dccc63074b74",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "anand\n"
     ]
    }
   ],
   "source": [
    "print(\"anand\");"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 15,
   "id": "59a6c71a-b9c7-409d-b4f2-a1f9c4dd81fe",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "Collecting openpyxl\n",
      "  Downloading openpyxl-3.1.3-py2.py3-none-any.whl (251 kB)\n",
      "\u001b[2K     \u001b[90m━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\u001b[0m \u001b[32m251.3/251.3 kB\u001b[0m \u001b[31m20.4 MB/s\u001b[0m eta \u001b[36m0:00:00\u001b[0m\n",
      "\u001b[?25hCollecting et-xmlfile (from openpyxl)\n",
      "  Downloading et_xmlfile-1.1.0-py3-none-any.whl (4.7 kB)\n",
      "Installing collected packages: et-xmlfile, openpyxl\n",
      "Successfully installed et-xmlfile-1.1.0 openpyxl-3.1.3\n"
     ]
    }
   ],
   "source": [
    "import requests # you need this module to make an API call\n",
    "import pandas as pd\n",
    "!pip install openpyxl"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 3,
   "id": "16dbbaf6-5ce5-4310-8a97-09f363f3a251",
   "metadata": {
    "tags": []
   },
   "outputs": [],
   "source": [
    "api_url = \"http://api.open-notify.org/astros.json\" "
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 5,
   "id": "05c73d35-17d5-4c0b-bbea-c9bca5fbc1d9",
   "metadata": {
    "tags": []
   },
   "outputs": [],
   "source": [
    "response = requests.get(api_url)"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 6,
   "id": "b45b321e-c6b1-4b03-ac1c-4e16776b38f0",
   "metadata": {
    "tags": []
   },
   "outputs": [],
   "source": [
    "if response.ok:             # if all is well() no errors, no network timeouts)\n",
    "    data = response.json()"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 7,
   "id": "3c1f82ec-c6b7-4281-b8b2-a3576fa8be27",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "{'people': [{'craft': 'ISS', 'name': 'Oleg Kononenko'}, {'craft': 'ISS', 'name': 'Nikolai Chub'}, {'craft': 'ISS', 'name': 'Tracy Caldwell Dyson'}, {'craft': 'ISS', 'name': 'Matthew Dominick'}, {'craft': 'ISS', 'name': 'Michael Barratt'}, {'craft': 'ISS', 'name': 'Jeanette Epps'}, {'craft': 'ISS', 'name': 'Alexander Grebenkin'}, {'craft': 'ISS', 'name': 'Butch Wilmore'}, {'craft': 'ISS', 'name': 'Sunita Williams'}, {'craft': 'Tiangong', 'name': 'Li Guangsu'}, {'craft': 'Tiangong', 'name': 'Li Cong'}, {'craft': 'Tiangong', 'name': 'Ye Guangfu'}], 'number': 12, 'message': 'success'}\n"
     ]
    }
   ],
   "source": [
    "print(data);"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 8,
   "id": "d071be68-5424-4367-967a-2be7fd3a9094",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "12\n"
     ]
    }
   ],
   "source": [
    "print(data.get('number'))\n"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 9,
   "id": "2e42cbbf-e70a-42eb-b920-baa3e1637da4",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "There are 12 astronauts on ISS\n",
      "And their names are :\n",
      "Oleg Kononenko\n",
      "Nikolai Chub\n",
      "Tracy Caldwell Dyson\n",
      "Matthew Dominick\n",
      "Michael Barratt\n",
      "Jeanette Epps\n",
      "Alexander Grebenkin\n",
      "Butch Wilmore\n",
      "Sunita Williams\n",
      "Li Guangsu\n",
      "Li Cong\n",
      "Ye Guangfu\n"
     ]
    }
   ],
   "source": [
    "astronauts = data.get('people')\n",
    "print(\"There are {} astronauts on ISS\".format(len(astronauts)))\n",
    "print(\"And their names are :\")\n",
    "for astronaut in astronauts:\n",
    "    print(astronaut.get('name'))"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 19,
   "id": "67f38b9d-5b64-4948-82a7-ad426476f100",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "Technology: Python, Number of Jobs: 100\n"
     ]
    }
   ],
   "source": [
    "import requests\n",
    "\n",
    "api_url = \"http://127.0.0.1:5000/data\"\n",
    "\n",
    "def get_number_of_jobs_T(technology):\n",
    "    # Make a GET request to the API with the technology as a parameter\n",
    "    response = requests.get(api_url, params={'technology': technology})\n",
    "    \n",
    "    # Check if the request was successful\n",
    "    if response.status_code == 200:\n",
    "        # Parse the JSON response\n",
    "        data = response.json()\n",
    "        \n",
    "        \n",
    "        # Extract the number of jobs from the response\n",
    "        # Default to 0 if the 'number_of_jobs' key is not found\n",
    "        number_of_jobs = data.get('number_of_jobs', 0)\n",
    "    else:\n",
    "        # If the request failed, set number_of_jobs to 0\n",
    "        number_of_jobs = 0\n",
    "    \n",
    "    # Return the technology and the number of jobs\n",
    "    return technology, number_of_jobs\n",
    "\n",
    "technology, number_of_jobs = get_number_of_jobs_T(\"Python\")\n",
    "print(f\"Technology: {technology}, Number of Jobs: {number_of_jobs}\")\n"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 11,
   "id": "2b2c1357-e865-4168-969c-2d1d02f8b1d1",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "here\n",
      "There are 100 jobs in Los Angeles.\n"
     ]
    }
   ],
   "source": [
    "api_url = \"http://127.0.0.1:5000/data\"\n",
    "\n",
    "def get_number_of_jobs_in_location(location):\n",
    "    \"\"\"\n",
    "    Function to get the number of jobs in the US for a given location.\n",
    "    \n",
    "    Parameters:\n",
    "    location (str): The location for which to find the number of jobs.\n",
    "    \n",
    "    Returns:\n",
    "    tuple: A tuple containing the location and the number of jobs.\n",
    "    \"\"\"\n",
    "    # Prepare the parameters to be sent in the request\n",
    "    params = {'location': location}\n",
    "\n",
    "    # Send a GET request to the API with the location as a parameter\n",
    "    response = requests.get(api_url, params=params)\n",
    "\n",
    "    # Check if the request was successful (status code 200)\n",
    "    if response.status_code == 200:\n",
    "        print(\"here\");\n",
    "        # Extract the number of jobs from the JSON response\n",
    "        data = response.json()\n",
    "        number_of_jobs = data.get('number_of_jobs', 0)  # Default to 0 if key is not found\n",
    "    else:\n",
    "        # If the request failed, return an error message or handle it accordingly\n",
    "        number_of_jobs = 0  # Default to 0 if the request fails\n",
    "        print(\"there\");\n",
    "    # Return the location and the number of jobs\n",
    "    return location, number_of_jobs\n",
    "\n",
    "# Example usage\n",
    "location = \"Los Angeles\"\n",
    "loc, jobs = get_number_of_jobs_in_location(location)\n",
    "print(f\"There are {jobs} jobs in {loc}.\")"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 12,
   "id": "175c2690-c381-4f05-b8d3-9d4eaec7f1e0",
   "metadata": {
    "tags": []
   },
   "outputs": [],
   "source": [
    "api_url = \"http://127.0.0.1:5000/data\"\n",
    "\n",
    "technologies = [\n",
    "    \"Python\",\n",
    "    \"Java\",\n",
    "    \"JavaScript\",\n",
    "    \"C#\",\n",
    "    \"C++\",\n",
    "    \"Ruby\",\n",
    "    \"Go\",\n",
    "    \"Swift\",\n",
    "    \"PHP\",\n",
    "    \"Kotlin\"\n",
    "]"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 16,
   "id": "73e94277-d984-4247-8b26-985e219021a3",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "Job postings data has been successfully written to job_postings.xlsx\n"
     ]
    }
   ],
   "source": [
    "import requests\n",
    "import pandas as pd\n",
    "\n",
    "api_url = \"http://127.0.0.1:5000/data\"\n",
    "\n",
    "technologies = [\n",
    "    \"Python\",\n",
    "    \"Java\",\n",
    "    \"JavaScript\",\n",
    "    \"C#\",\n",
    "    \"C++\",\n",
    "    \"Ruby\",\n",
    "    \"Go\",\n",
    "    \"Swift\",\n",
    "    \"PHP\",\n",
    "    \"Kotlin\"\n",
    "]\n",
    "\n",
    "def get_number_of_jobs(technology):\n",
    "    try:\n",
    "        response = requests.get(api_url, params={'technology': technology})\n",
    "        response.raise_for_status()  # Raise an error for bad responses\n",
    "        data = response.json()\n",
    "        number_of_jobs = data.get('number_of_jobs', 0)\n",
    "    except requests.exceptions.RequestException as e:\n",
    "        print(f\"An error occurred while fetching data for {technology}: {e}\")\n",
    "        number_of_jobs = None\n",
    "    return number_of_jobs\n",
    "\n",
    "job_data = []\n",
    "\n",
    "for tech in technologies:\n",
    "    jobs = get_number_of_jobs(tech)\n",
    "    job_data.append({'Technology': tech, 'Number of Jobs': jobs})\n",
    "\n",
    "df = pd.DataFrame(job_data)\n",
    "\n",
    "excel_file = 'job_postings.xlsx'\n",
    "df.to_excel(excel_file, index=False)\n",
    "\n",
    "print(f\"Job postings data has been successfully written to {excel_file}\")"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 20,
   "id": "22a6f555-b4cc-4c8c-a38a-a9e283b650aa",
   "metadata": {
    "tags": []
   },
   "outputs": [],
   "source": [
    "#Create a WorkBook"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 23,
   "id": "5bd73d9d-7554-4ce9-987a-69ca2c28917c",
   "metadata": {
    "tags": []
   },
   "outputs": [],
   "source": [
    "from openpyxl import Workbook\n",
    "\n",
    "# Create a new workbook\n",
    "wb = Workbook()\n",
    "\n",
    "# Select the active worksheet\n",
    "ws = wb.active\n",
    "\n",
    "# Optionally, rename the active worksheet\n",
    "ws.title = \"Sheet1\"\n",
    "\n",
    "# Add some sample data\n",
    "ws['A1'] = 'Technology'\n",
    "ws['B1'] = 'Number of Jobs'\n",
    "\n",
    "# Sample data\n",
    "technologies = ['Python', 'Java', 'JavaScript']\n",
    "job_counts = [1200, 950, 1100]\n",
    "\n",
    "for index, (tech, count) in enumerate(zip(technologies, job_counts), start=2):\n",
    "    ws[f'A{index}'] = tech\n",
    "    ws[f'B{index}'] = count\n",
    "\n",
    "# Save the workbook to a file\n",
    "wb.save('my_workbook.xlsx')"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 24,
   "id": "d66337c5-d0ff-48fc-8d55-b22f23816b70",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "Workbook created and saved as 'job_postings.xlsx'\n"
     ]
    }
   ],
   "source": [
    "#Find the number of jobs postings for each of the technology in the above list. Write the technology name and the number of jobs postings into the excel spreadsheet\n",
    "\n",
    "# Fetch job postings for each technology and write to the Excel file\n",
    "\n",
    "for index, tech in enumerate(technologies, start=2):\n",
    "    number_of_jobs = get_number_of_jobs(tech)\n",
    "    ws[f'A{index}'] = tech\n",
    "    ws[f'B{index}'] = number_of_jobs\n",
    "\n",
    "# Save the workbook to a file\n",
    "wb.save('job_postings.xlsx')\n",
    "\n",
    "print(\"Workbook created and saved as 'job_postings.xlsx'\")"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 25,
   "id": "8f70a345-76cb-4632-9638-b33824ff34cd",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "Workbook created and saved as 'github-job-postings.xlsx'\n"
     ]
    }
   ],
   "source": [
    "# Save the workbook to a file named 'github-job-postings.xlsx'\n",
    "wb.save('github-job-postings.xlsx')\n",
    "\n",
    "print(\"Workbook created and saved as 'github-job-postings.xlsx'\")"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": 27,
   "id": "c224ac7e-9ec9-4e56-a074-7f6b201713e7",
   "metadata": {
    "tags": []
   },
   "outputs": [
    {
     "name": "stdout",
     "output_type": "stream",
     "text": [
      "Workbook created and saved as 'github-job-postings.xlsx'\n"
     ]
    }
   ],
   "source": [
    "#Collect the number of job postings for the following languages using the API:\n",
    "api_url = \"http://127.0.0.1:5000/data\"\n",
    "\n",
    "# List of technologies to query\n",
    "technologies = [\n",
    "    \"C\",\n",
    "    \"C#\",\n",
    "    \"C++\",\n",
    "    \"Java\",\n",
    "    \"JavaScript\",\n",
    "    \"Python\",\n",
    "    \"Scala\",\n",
    "    \"Oracle\",\n",
    "    \"SQL Server\",\n",
    "    \"MySQL Server\",\n",
    "    \"PostgreSQL\",\n",
    "    \"MongoDB\"\n",
    "]\n",
    "\n",
    "def get_number_of_jobs(technology):\n",
    "    try:\n",
    "        response = requests.get(api_url, params={'technology': technology})\n",
    "        response.raise_for_status()  # Raise an error for bad responses\n",
    "        data = response.json()\n",
    "        number_of_jobs = data.get('number_of_jobs', 0)\n",
    "    except requests.exceptions.RequestException as e:\n",
    "        print(f\"An error occurred while fetching data for {technology}: {e}\")\n",
    "        number_of_jobs = None\n",
    "    return number_of_jobs\n",
    "\n",
    "# Create a new workbook\n",
    "wb = Workbook()\n",
    "ws = wb.active\n",
    "ws.title = \"Job Postings\"\n",
    "\n",
    "# Add headers\n",
    "ws['A1'] = 'Technology'\n",
    "ws['B1'] = 'Number of Jobs'\n",
    "\n",
    "# Fetch job postings for each technology and write to the Excel file\n",
    "for index, tech in enumerate(technologies, start=2):\n",
    "    number_of_jobs = get_number_of_jobs(tech)\n",
    "    ws[f'A{index}'] = tech\n",
    "    ws[f'B{index}'] = number_of_jobs\n",
    "\n",
    "# Save the workbook to a file named 'github-job-postings.xlsx'\n",
    "wb.save('github-job-postings.xlsx')\n",
    "\n",
    "print(\"Workbook created and saved as 'github-job-postings.xlsx'\")"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": null,
   "id": "178b3c51-a423-45a9-b768-e838bbc2e443",
   "metadata": {},
   "outputs": [],
   "source": []
  }
 ],
 "metadata": {
  "kernelspec": {
   "display_name": "Python",
   "language": "python",
   "name": "conda-env-python-py"
  },
  "language_info": {
   "codemirror_mode": {
    "name": "ipython",
    "version": 3
   },
   "file_extension": ".py",
   "mimetype": "text/x-python",
   "name": "python",
   "nbconvert_exporter": "python",
   "pygments_lexer": "ipython3",
   "version": "3.7.12"
  }
 },
 "nbformat": 4,
 "nbformat_minor": 5
}
